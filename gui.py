"""
学生作业文件批量命名规范化工具 —— GUI 界面 (PySide6)

功能：图形化配置 → 预览匹配结果 → 执行重命名（带进度条），
      与 CLI 模式（main.py）共用 core/ 模块和 config.json。
"""

import json
import os
import sys

from PySide6.QtCore import Qt, QThread, Signal
import subprocess

from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenu,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSlider,
    QSpinBox,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

# 复用已有核心模块
from core.matcher import build_pinyin_map, match_files
from core.renamer import undo_renaming
from core.reporter import export_results
from core.checker import check_submissions

# 配置文件路径（与可执行文件同目录，兼容开发与打包两种模式）
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(sys.executable)      # 打包后：exe 所在目录
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))  # 开发时：脚本所在目录
DEFAULT_CONFIG = os.path.join(SCRIPT_DIR, "config.json")


# ═══════════════════════════════════════════════════════════════════════
# 后台工作线程
# ═══════════════════════════════════════════════════════════════════════

class RenameWorker(QThread):
    """后台执行重命名操作，逐文件发射进度信号，并写入操作日志。"""
    progress    = Signal(int, str)
    file_done   = Signal(str, str, bool, str)
    all_done    = Signal(int, int)

    def __init__(self, results, folder_path, log_path=None):
        super().__init__()
        self.results     = results
        self.folder_path = folder_path
        self.log_path    = log_path
        self.log_entries = []

    def run(self):
        import csv
        from datetime import datetime

        renamed = 0
        skipped = 0
        matched = [r for r in self.results if r["状态"] == "✅ 已匹配"]

        for i, r in enumerate(matched):
            old_name = r["原文件名"]
            new_name = r["建议新文件名"]
            old_path = os.path.join(self.folder_path, old_name)
            new_path = os.path.join(self.folder_path, new_name)
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            self.progress.emit(i + 1, f"正在处理: {old_name}")

            try:
                if old_name == new_name:
                    self.file_done.emit(old_name, new_name, False, "新旧文件名相同")
                    self._log(timestamp, old_name, new_name, "跳过", "新旧文件名相同")
                    skipped += 1
                    continue
                if not os.path.isfile(old_path):
                    self.file_done.emit(old_name, new_name, False, "原文件不存在")
                    self._log(timestamp, old_name, new_name, "跳过", "原文件不存在")
                    skipped += 1
                    continue
                if os.path.exists(new_path):
                    self.file_done.emit(old_name, new_name, False, "目标已存在")
                    self._log(timestamp, old_name, new_name, "跳过", "目标已存在")
                    skipped += 1
                    continue

                os.rename(old_path, new_path)
                self.file_done.emit(old_name, new_name, True, "")
                self._log(timestamp, old_name, new_name, "成功", "")
                renamed += 1
            except Exception as e:
                self.file_done.emit(old_name, new_name, False, str(e))
                self._log(timestamp, old_name, new_name, "跳过", str(e))
                skipped += 1

        # 保存日志
        if self.log_path and self.log_entries:
            self._write_log()

        self.all_done.emit(renamed, skipped)

    def _log(self, timestamp, old_name, new_name, status, reason):
        self.log_entries.append({
            "时间": timestamp, "原文件名": old_name,
            "新文件名": new_name, "状态": status, "备注": reason,
        })

    def _write_log(self):
        import csv
        log_dir = os.path.dirname(self.log_path)
        if log_dir:
            os.makedirs(log_dir, exist_ok=True)
        with open(self.log_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=["时间", "原文件名", "新文件名", "状态", "备注"])
            w.writeheader()
            w.writerows(self.log_entries)


class MatchWorker(QThread):
    """后台执行匹配计算，完成后返回结果。"""
    finished = Signal(list)  # results list

    def __init__(self, filenames, name_id_map, pinyin_map, template, threshold, enable_pinyin, extract_keyword):
        super().__init__()
        self.filenames       = filenames
        self.name_id_map     = name_id_map
        self.pinyin_map      = pinyin_map
        self.template        = template
        self.threshold       = threshold
        self.enable_pinyin   = enable_pinyin
        self.extract_keyword = extract_keyword

    def run(self):
        results = match_files(
            self.filenames, self.name_id_map, self.pinyin_map,
            self.template, self.threshold, self.enable_pinyin,
            self.extract_keyword,
        )
        self.finished.emit(results)


# ═══════════════════════════════════════════════════════════════════════
# 拖拽输入框
# ═══════════════════════════════════════════════════════════════════════

class DragDropLineEdit(QLineEdit):
    """支持拖拽文件/文件夹的输入框。

    拖拽模式：
        - "folder": 仅接受文件夹
        - "file":   仅接受 Excel 文件 (.xlsx/.xls)
    """

    def __init__(self, mode="folder", parent=None):
        super().__init__(parent)
        self._mode = mode
        self.setAcceptDrops(True)
        self._highlight = False

    def dragEnterEvent(self, event):
        if self._accept(event):
            self._highlight = True
            self.setStyleSheet(
                "DragDropLineEdit { border: 2px solid #1565c0;"
                "background-color: #e3f2fd; border-radius: 4px; padding: 10px 8px; }"
            )
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if self._highlight:
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        self._highlight = False
        self.setStyleSheet("")
        event.accept()

    def dropEvent(self, event):
        self._highlight = False
        self.setStyleSheet("")
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                path = urls[0].toLocalFile()
                if self._is_valid(path):
                    self.setText(path)
                    event.acceptProposedAction()
                    return
        event.ignore()

    def _accept(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if len(urls) == 1:
                return self._is_valid(urls[0].toLocalFile())
        return False

    def _is_valid(self, path):
        """校验拖入路径是否有效。"""
        if self._mode == "folder":
            return os.path.isdir(path)
        else:
            return os.path.isfile(path) and path.lower().endswith((".xlsx", ".xls"))


# ═══════════════════════════════════════════════════════════════════════
# 列映射对话框
# ═══════════════════════════════════════════════════════════════════════

class ColumnMappingDialog(QDialog):
    """Excel 列映射向导：当首行缺少「姓名」/「学号」时，让用户手动选择对应列。"""

    def __init__(self, headers, parent=None):
        super().__init__(parent)
        self.setWindowTitle("列映射向导")
        self.resize(400, 180)
        self._result = None  # (col_name_index, col_id_index) or None

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # 提示文字
        hint = QLabel("Excel 表头未自动匹配到「姓名」或「学号」列，请手动指定：")
        hint.setWordWrap(True)
        layout.addWidget(hint)

        # 姓名列选择
        form = QFormLayout()
        self.cb_name = QComboBox()
        self.cb_name.addItems(headers)
        self._auto_select(self.cb_name, ["姓名", "名字", "学生姓名", "name"])
        form.addRow("姓名列：", self.cb_name)

        # 学号列选择
        self.cb_id = QComboBox()
        self.cb_id.addItems(headers)
        self._auto_select(self.cb_id, ["学号", "工号", "学生学号", "id", "编号"])
        form.addRow("学号列：", self.cb_id)
        layout.addLayout(form)

        # 按钮
        btn_row = QHBoxLayout()
        btn_ok = QPushButton("确定")
        btn_ok.clicked.connect(self._on_ok)
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

    def _auto_select(self, combo, keywords):
        """自动选中第一个包含关键词的选项。"""
        for i in range(combo.count()):
            text = combo.itemText(i).lower()
            if any(kw.lower() in text for kw in keywords):
                combo.setCurrentIndex(i)
                return

    def _on_ok(self):
        self._result = (self.cb_name.currentIndex(), self.cb_id.currentIndex())
        self.accept()

    def get_mapping(self):
        return self._result


# ═══════════════════════════════════════════════════════════════════════
# 主窗口
# ═══════════════════════════════════════════════════════════════════════

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("学生作业文件批量命名规范化工具")
        self.resize(1050, 700)
        self.setAcceptDrops(True)   # 窗口级拖拽作为备用

        self._results = []          # 最近一次匹配结果
        self._name_id_map = {}      # 姓名→学号映射
        self._pinyin_map = {}       # 拼音映射
        self._last_log_path = ""    # 最近一次重命名的日志路径

        self._setup_ui()
        self._apply_stylesheet()
        self._load_config()

    # ── 窗口级拖拽（备用路由） ──────────────────────────────────────────

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls() and len(event.mimeData().urls()) == 1:
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        event.acceptProposedAction()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                path = urls[0].toLocalFile()
                if os.path.isdir(path):
                    self.le_folder.setText(path)
                elif os.path.isfile(path) and path.lower().endswith((".xlsx", ".xls")):
                    self.le_excel.setText(path)

    # ── QSS 样式 ──────────────────────────────────────────────────────

    def _apply_stylesheet(self):
        self.setStyleSheet("""
            /* 主背景：清新的高明度莫兰迪蓝渐变 */
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                            stop:0 #e0eafc, stop:1 #cfdef3);
            }

            /* 卡片：纯净的半透明白，保证黑字在上面阅读舒适 */
            QGroupBox {
                background-color: rgba(255, 255, 255, 0.85);
                border: 1px solid rgba(255, 255, 255, 0.5);
                border-radius: 16px;
                margin-top: 15px;
                color: #1e293b;
                font-weight: 700;
                padding-top: 20px;
            }

            /* 输入框：白色实底，确保文字清晰可见 */
            QLineEdit {
                background-color: #ffffff;
                border: 1px solid #cbd5e1;
                border-radius: 6px;
                color: #0f172a;
                padding: 8px;
            }
            QLineEdit:focus {
                border: 2px solid #60a5fa;
            }
            /* SpinBox：留空，避免 QSS 破坏原生箭头 */

            /* 按钮共享基础 — 仅作用于带 objectName 的按钮，避免污染 SpinBox 箭头 */
            QPushButton#btnPreview, QPushButton#btnApply, QPushButton#btnUndo,
            QPushButton#btnSave, QPushButton#btnBrowse, QPushButton#btnToggle {
                border-radius: 6px;
                color: white;
                font-weight: 600;
                padding: 8px 16px;
                border: none;
            }
            QPushButton#btnPreview { background-color: #7a8f9e; }
            QPushButton#btnApply  { background-color: #6d947a; }
            QPushButton#btnUndo   { background-color: #c8967a; }

            /* 表格：清晰的黑字阅读模式 */
            QTableWidget {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                color: #1e293b;
                gridline-color: #dde1e6;
            }
            QHeaderView::section {
                background-color: #f1f5f9;
                color: #475569;
                font-weight: bold;
                border: none;
                border-right: 1px solid #cbd5e1;
                padding: 8px 12px;
            }

            /* 保存按钮（莫兰迪灰紫） */
            QPushButton#btnSave { background-color: #9b8da5; }

            /* 浏览按钮（莫兰迪灰） */
            QPushButton#btnBrowse {
                background-color: #8a95a0;
                padding: 6px 12px;
                font-size: 12px;
            }

            /* 切换标签按钮 */
            QPushButton#btnToggle {
                background-color: rgba(255,255,255,0.6);
                color: #475569;
                border: 1px solid #cbd5e1;
                font-size: 13px;
                font-weight: 600;
                padding: 8px 24px;
                border-radius: 0px;
            }
            QPushButton#btnToggle:first-child {
                border-top-left-radius: 8px;
                border-bottom-left-radius: 8px;
            }
            QPushButton#btnToggle:last-child {
                border-top-right-radius: 8px;
                border-bottom-right-radius: 8px;
            }
            QPushButton#btnToggle:checked {
                background-color: #7a8f9e;
                color: #ffffff;
                border-color: #475569;
            }
            QPushButton#btnToggle:hover:!checked {
                background-color: rgba(255,255,255,0.9);
            }

            /* 进度条 */
            QProgressBar {
                border: 1px solid #cbd5e1;
                border-radius: 6px;
                text-align: center;
                height: 20px;
                background-color: #f1f5f9;
                color: #475569;
                font-weight: 600;
            }
            QProgressBar::chunk {
                background-color: #6d947a;
                border-radius: 5px;
            }

            /* 滑块 */
            QSlider {
                min-height: 30px;
            }
            QSlider::groove:horizontal {
                height: 6px;
                background: #e2e8f0;
                border-radius: 3px;
            }
            QSlider::handle:horizontal {
                width: 16px;
                height: 16px;
                margin: -5px 0;
                background: #7a8f9e;
                border-radius: 8px;
            }

            /* 复选框 */
            QCheckBox {
                color: #475569;
                font-size: 13px;
                spacing: 8px;
            }

            /* 状态栏 */
            QStatusBar {
                color: #94a3b8;
                font-size: 12px;
            }

            /* 拖拽输入框：虚线提示 */
            DragDropLineEdit {
                border: 2px dashed #bccbd8;
                padding: 10px 10px;
                font-size: 13px;
                background-color: rgba(255,255,255,0.7);
                border-radius: 6px;
            }
            DragDropLineEdit:hover {
                border-color: #7a8f9e;
                background-color: rgba(255,255,255,0.95);
            }
        """)
    

    # ── UI 构建 ──────────────────────────────────────────────────────

    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setSpacing(12)

        # ── 配置区 ──
        gb_config = QGroupBox("配置")
        form = QFormLayout(gb_config)
        form.setSpacing(8)

        # 文件夹（支持拖拽）
        row_folder = QHBoxLayout()
        self.le_folder = DragDropLineEdit(mode="folder")
        self.le_folder.setPlaceholderText("请拖入文件夹或点击浏览选择文件夹")
        btn_folder = QPushButton("📂 浏览...")
        btn_folder.setObjectName("btnBrowse")
        btn_folder.clicked.connect(self._browse_folder)
        row_folder.addWidget(self.le_folder)
        row_folder.addWidget(btn_folder)
        form.addRow("文件夹：", row_folder)

        # Excel（支持拖拽）
        row_excel = QHBoxLayout()
        self.le_excel = DragDropLineEdit(mode="file")
        self.le_excel.setPlaceholderText("请拖入名单或点击浏览选择名单（支持 .xlsx / .xls）")
        btn_excel = QPushButton("📄 浏览...")
        btn_excel.setObjectName("btnBrowse")
        btn_excel.clicked.connect(self._browse_excel)
        row_excel.addWidget(self.le_excel)
        row_excel.addWidget(btn_excel)
        form.addRow("名单：", row_excel)

        # 命名模板
        self.le_template = QLineEdit()
        self.le_template.setPlaceholderText("{学号}{姓名}{匹配项}（默认，可自行修改）")
        form.addRow("命名模板：", self.le_template)

        # 匹配项（用于 {匹配项} 占位符）
        self.le_keyword = QLineEdit()
        self.le_keyword.setPlaceholderText("如输入\"实验3\"，可在文件名中匹配 part3/lab3 等并统一输出\"实验3\"")
        form.addRow("匹配项：", self.le_keyword)

        # 匹配阈值
        row_threshold = QHBoxLayout()
        self.sb_threshold = QSpinBox()
        self.sb_threshold.setRange(0, 100)
        self.sb_threshold.setValue(80)
        self.sb_threshold.setSuffix("%")
        self.slider_threshold = QSlider()
        self.slider_threshold.setOrientation(Qt.Orientation.Horizontal)
        self.slider_threshold.setRange(0, 100)
        self.slider_threshold.setValue(80)
        self.sb_threshold.valueChanged.connect(self.slider_threshold.setValue)
        self.slider_threshold.valueChanged.connect(self.sb_threshold.setValue)
        row_threshold.addWidget(self.slider_threshold)
        row_threshold.addWidget(self.sb_threshold)
        form.addRow("匹配阈值：", row_threshold)

        # 拼音开关
        self.cb_pinyin = QCheckBox("启用拼音匹配（zhangsan / zs）")
        self.cb_pinyin.setChecked(True)
        form.addRow("", self.cb_pinyin)

        root.addWidget(gb_config)

        # ── 操作按钮 ──
        row_btn = QHBoxLayout()
        row_btn.setSpacing(12)

        self.btn_preview = QPushButton(" 预览匹配结果")
        self.btn_preview.setObjectName("btnPreview")
        self.btn_preview.setMinimumHeight(42)
        self.btn_preview.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_preview.clicked.connect(self._on_preview)

        self.btn_apply = QPushButton(" 执行重命名")
        self.btn_apply.setObjectName("btnApply")
        self.btn_apply.setMinimumHeight(42)
        self.btn_apply.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_apply.clicked.connect(self._on_apply)

        self.btn_save_config = QPushButton(" 保存配置")
        self.btn_save_config.setObjectName("btnSave")
        self.btn_save_config.setMinimumHeight(42)
        self.btn_save_config.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_save_config.clicked.connect(self._save_config)

        self.btn_undo = QPushButton(" 撤销上次重命名")
        self.btn_undo.setObjectName("btnUndo")
        self.btn_undo.setMinimumHeight(42)
        self.btn_undo.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_undo.clicked.connect(self._on_undo)

        row_btn.addWidget(self.btn_preview)
        row_btn.addWidget(self.btn_apply)
        row_btn.addWidget(self.btn_save_config)
        row_btn.addWidget(self.btn_undo)
        row_btn.addStretch()
        root.addLayout(row_btn)

        # ── 进度条 ──
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        root.addWidget(self.progress)

        # ── 结果区域（切换按钮 + 表格） ──
        # 切换按钮
        row_toggle = QHBoxLayout()
        row_toggle.setSpacing(0)

        self.btn_show_match = QPushButton("📋 匹配结果")
        self.btn_show_match.setObjectName("btnToggle")
        self.btn_show_match.setCheckable(True)
        self.btn_show_match.setChecked(True)
        self.btn_show_match.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_show_match.clicked.connect(lambda: self._toggle_view(0))

        self.btn_show_check = QPushButton("📊 异常情况")
        self.btn_show_check.setObjectName("btnToggle")
        self.btn_show_check.setCheckable(True)
        self.btn_show_check.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_show_check.clicked.connect(lambda: self._toggle_view(1))

        row_toggle.addWidget(self.btn_show_match)
        row_toggle.addWidget(self.btn_show_check)
        row_toggle.addStretch()
        root.addLayout(row_toggle)

        # 堆叠视图
        self.stack = QStackedWidget()

        # 页面0：匹配结果
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels([
            "状态", "原文件名", "建议新文件名", "匹配分数", "匹配方式",
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setColumnWidth(0, 90)
        self.table.setColumnWidth(1, 180)
        self.table.setColumnWidth(2, 180)
        self.table.setColumnWidth(3, 80)
        self.table.setEditTriggers(QTableWidget.DoubleClicked)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._on_context_menu)
        self.table.cellChanged.connect(self._on_table_cell_changed)
        self.stack.addWidget(self.table)

        # 页面1：异常情况
        self.check_table = QTableWidget()
        self.check_table.setColumnCount(5)
        self.check_table.setHorizontalHeaderLabels([
            "姓名", "学号", "提交数", "状态", "匹配文件",
        ])
        self.check_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.check_table.horizontalHeader().setStretchLastSection(True)
        self.check_table.setColumnWidth(0, 80)
        self.check_table.setColumnWidth(1, 100)
        self.check_table.setColumnWidth(2, 60)
        self.check_table.setColumnWidth(3, 90)
        self.check_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.check_table.setAlternatingRowColors(True)
        self.check_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.check_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.check_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.check_table.customContextMenuRequested.connect(self._on_check_context_menu)
        self.stack.addWidget(self.check_table)

        root.addWidget(self.stack, stretch=1)

        # ── 状态栏 ──
        self.statusBar().showMessage("就绪")

    # ── 配置读写 ──────────────────────────────────────────────────────

    def _load_config(self):
        """从 config.json 加载配置并填入界面。"""
        if not os.path.exists(DEFAULT_CONFIG):
            self.statusBar().showMessage("未找到 config.json，使用默认值")
            return

        try:
            with open(DEFAULT_CONFIG, "r", encoding="utf-8") as f:
                cfg = json.load(f)

            self.le_folder.setText(cfg.get("folder_path", ""))
            self.le_excel.setText(cfg.get("excel_path", ""))
            self.le_template.setText(cfg.get("template", "{学号}{姓名}{匹配项}"))
            self.le_keyword.setText(cfg.get("extract_keyword", ""))
            self.sb_threshold.setValue(cfg.get("match_threshold", 80))
            self.cb_pinyin.setChecked(cfg.get("enable_pinyin", True))

            self.statusBar().showMessage(f"配置已加载 —— {DEFAULT_CONFIG}")
        except Exception as e:
            self.statusBar().showMessage(f"配置加载失败：{e}")

    def _save_config(self):
        """将界面当前值保存到 config.json。"""
        # 校验模板非法字符
        illegal = set(r'\/:*?"<>|')
        template = self.le_template.text().strip() or "{学号}{姓名}{匹配项}"
        if illegal & set(template):
            QMessageBox.warning(self, "模板错误",
                f"模板包含非法字符：{''.join(illegal & set(template))}")
            return

        cfg = {
            "folder_path":     self.le_folder.text().strip(),
            "excel_path":      self.le_excel.text().strip(),
            "output_path":     "match_result.csv",
            "template":        template,
            "match_threshold": self.sb_threshold.value(),
            "enable_pinyin":   self.cb_pinyin.isChecked(),
            "extract_keyword": self.le_keyword.text().strip(),
        }

        try:
            with open(DEFAULT_CONFIG, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=4)
            self.statusBar().showMessage(f"配置已保存 —— {DEFAULT_CONFIG}")
        except Exception as e:
            QMessageBox.critical(self, "保存失败", str(e))

    # ── 文件浏览 ─────────────────────────────────────────────────────

    def _browse_folder(self):
        path = QFileDialog.getExistingDirectory(self, "选择文件夹")
        if path:
            self.le_folder.setText(path)

    def _browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "",
            "Excel 文件 (*.xlsx *.xls);;所有文件 (*)"
        )
        if path:
            self.le_excel.setText(path)

    # ── 输入校验 ─────────────────────────────────────────────────────

    def _read_excel_headers_rows(self, excel_path):
        """读取 Excel 的 headers 和 rows，兼容 .xlsx 和 .xls 格式。

        Returns:
            (headers: list[str], rows: list[list]) 或 (None, None) 失败时
        """
        import openpyxl

        if excel_path.lower().endswith(".xls") and not excel_path.lower().endswith(".xlsx"):
            # 旧格式 .xls → 使用 xlrd
            try:
                import xlrd
                wb = xlrd.open_workbook(excel_path)
                ws = wb.sheet_by_index(0)
                headers = [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
                rows = [
                    [ws.cell_value(r, c) for c in range(ws.ncols)]
                    for r in range(1, ws.nrows)
                ]
                return headers, rows
            except Exception as e:
                QMessageBox.critical(self, "读取失败", f"无法读取 .xls 文件：{e}")
                return None, None
        else:
            # 新格式 .xlsx → 使用 openpyxl
            try:
                wb = openpyxl.load_workbook(excel_path, read_only=True)
                ws = wb.active
                headers = [
                    str(c.value) if c.value is not None else ""
                    for c in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                rows = [
                    [c for c in row]
                    for row in ws.iter_rows(min_row=2, values_only=True)
                ]
                wb.close()
                return headers, rows
            except Exception as e:
                QMessageBox.critical(self, "读取失败", f"无法读取 Excel：{e}")
                return None, None

    def _validate_inputs(self):
        """校验界面输入，返回 (filenames, name_id_map) 或 (None, None)。"""
        folder_path = self.le_folder.text().strip()
        excel_path  = self.le_excel.text().strip()

        if not folder_path or not os.path.isdir(folder_path):
            QMessageBox.warning(self, "输入错误", "请选择有效的文件夹路径。")
            return None, None
        if not excel_path or not os.path.isfile(excel_path):
            QMessageBox.warning(self, "输入错误", "请选择有效的 Excel 文件。")
            return None, None

        # 读取 Excel（兼容 .xls / .xlsx）
        headers, rows = self._read_excel_headers_rows(excel_path)
        if headers is None:
            return None, None

        # 自动匹配或弹窗让用户选择
        if "姓名" not in headers or "学号" not in headers:
            dlg = ColumnMappingDialog(headers, self)
            if dlg.exec() != QDialog.DialogCode.Accepted:
                return None, None
            mapping = dlg.get_mapping()
            if mapping is None:
                return None, None
            col_name, col_id = mapping
        else:
            col_name = headers.index("姓名")
            col_id   = headers.index("学号")

        # 逐行读取，跳过空值
        name_id_map = {}
        for row in rows:
            name = row[col_name]
            sid  = row[col_id]
            if name and sid:
                name = str(name).strip()
                sid  = str(sid).strip()
                if name and sid:
                    name_id_map[name] = sid

        if not name_id_map:
            QMessageBox.warning(self, "数据为空", "Excel 中没有有效的姓名-学号数据。")
            return None, None

        filenames = [
            f for f in os.listdir(folder_path)
            if os.path.isfile(os.path.join(folder_path, f))
        ]
        if not filenames:
            QMessageBox.warning(self, "文件夹为空", "未找到任何文件。")
            return None, None

        return filenames, name_id_map

    # ── 预览匹配 ─────────────────────────────────────────────────────

    def _on_preview(self):
        """点击「预览匹配结果」按钮。"""
        filenames, name_id_map = self._validate_inputs()
        if not filenames:
            return

        self._name_id_map = name_id_map
        self.btn_preview.setEnabled(False)
        self.btn_apply.setEnabled(False)
        self.statusBar().showMessage("正在匹配...")

        # 构建拼音映射
        enable_pinyin = self.cb_pinyin.isChecked()
        if enable_pinyin:
            self._pinyin_map = build_pinyin_map(name_id_map)
        else:
            self._pinyin_map = {}

        template  = self.le_template.text().strip() or "{学号}{姓名}{匹配项}"
        threshold = self.sb_threshold.value()

        extract_keyword = self.le_keyword.text().strip()

        self._match_worker = MatchWorker(
            filenames, name_id_map, self._pinyin_map,
            template, threshold, enable_pinyin, extract_keyword,
        )
        self._match_worker.finished.connect(self._on_match_done)
        self._match_worker.start()

    def _on_match_done(self, results):
        """匹配完成后排序 → 填充表格 → 异常情况检查。"""
        # 排序：❌ 未匹配 → ⚠️ 多重匹配 → ✅ 已匹配
        status_order = {"✅ 已匹配": 0, "⚠️ 多重匹配": 1, "❌ 未匹配": 2}
        results.sort(key=lambda r: status_order.get(r["状态"], 99))
        self._results = results
        self._populate_table(results)
        self._populate_check_table(results)

        # 导出 CSV
        output_path = "match_result.csv"
        export_results(results, output_path)

        self.btn_preview.setEnabled(True)
        self.btn_apply.setEnabled(True)

        matched   = sum(1 for r in results if r["状态"] == "✅ 已匹配")
        unmatched = sum(1 for r in results if r["状态"] == "❌ 未匹配")
        multi     = sum(1 for r in results if r["状态"] == "⚠️ 多重匹配")

        # 异常情况统计
        if self._name_id_map:
            report = check_submissions(results, self._name_id_map)
            normal  = sum(1 for r in report if r["状态"] == "✅ 正常")
            missing = sum(1 for r in report if r["状态"] == "❌ 缺交")
            extra   = sum(1 for r in report if r["状态"] == "⚠️ 多交")
            self.statusBar().showMessage(
                f"匹配完成 —— ✅ {matched}  |  ❌ {unmatched}  |  ⚠️ {multi}"
                f"  ||  异常情况 —— 正常 {normal} | 缺交 {missing} | 多交 {extra}"
            )
        else:
            self.statusBar().showMessage(
                f"匹配完成 —— ✅ {matched}  |  ❌ {unmatched}  |  ⚠️ {multi}"
            )

    def _populate_table(self, results):
        """将匹配结果填入 QTableWidget，并按状态标记行颜色。
        列0(状态)和列2(建议新文件名)可双击编辑，其余只读。
        """
        self._populating = True
        self.table.setRowCount(len(results))
        for row, r in enumerate(results):
            status = r["状态"]

            # 创建表格项
            item_status   = QTableWidgetItem(status)
            item_original = QTableWidgetItem(r["原文件名"])
            item_new      = QTableWidgetItem(r["建议新文件名"])
            item_score    = QTableWidgetItem(str(r["匹配分数"]))
            item_method   = QTableWidgetItem(r["匹配方式"])

            # 设置可编辑权限：列0(状态) 和 列2(建议新文件名) 可编辑
            flags_default = item_original.flags()  # 默认只读
            item_original.setFlags(flags_default & ~Qt.ItemFlag.ItemIsEditable)
            item_score.setFlags(flags_default & ~Qt.ItemFlag.ItemIsEditable)
            item_method.setFlags(flags_default & ~Qt.ItemFlag.ItemIsEditable)
            # item_status 和 item_new 保持默认（可编辑）

            # 根据状态设定行背景色
            if status == "✅ 已匹配":
                bg = QColor("#e8f5e9")  # 浅绿色
                fg = QColor("#2e7d32")  # 深绿字
            elif status == "❌ 未匹配":
                bg = QColor("#ffebee")  # 浅红色
                fg = QColor("#c62828")  # 深红字
            elif status == "⚠️ 多重匹配":
                bg = QColor("#fff8e1")  # 浅黄色
                fg = QColor("#e65100")  # 深橙字
            else:
                bg = QColor("#ffffff")
                fg = QColor("#333333")

            for item in (item_status, item_original, item_new, item_score, item_method):
                item.setBackground(bg)
                item.setForeground(fg)

            self.table.setItem(row, 0, item_status)
            self.table.setItem(row, 1, item_original)
            self.table.setItem(row, 2, item_new)
            self.table.setItem(row, 3, item_score)
            self.table.setItem(row, 4, item_method)

        self._populating = False

    def _on_table_cell_changed(self, row, col):
        """用户编辑表格后，同步回 _results。"""
        if getattr(self, '_populating', False):
            return  # 正在填充表格，忽略
        if not self._results or row >= len(self._results):
            return
        new_value = self.table.item(row, col).text().strip()
        if col == 0:  # 状态列
            self._results[row]["状态"] = new_value
        elif col == 2:  # 建议新文件名列
            self._results[row]["建议新文件名"] = new_value

    # ── 视图切换 ──────────────────────────────────────────────────────

    def _toggle_view(self, index):
        """切换结果视图：0=匹配结果，1=异常情况。"""
        self.stack.setCurrentIndex(index)
        self.btn_show_match.setChecked(index == 0)
        self.btn_show_check.setChecked(index == 1)

    # ── 右键菜单 ──────────────────────────────────────────────────────

    def _on_context_menu(self, pos):
        """在鼠标位置弹出右键菜单。"""
        rows = set(idx.row() for idx in self.table.selectedIndexes())
        if not rows:
            return

        menu = QMenu(self)
        single_row = (len(rows) == 1)
        row = min(rows)

        if single_row:
            r = self._results[row]
            old_name = r["原文件名"]
            new_name = r["建议新文件名"]
            status   = r["状态"]

            # 编辑
            act_edit = menu.addAction("📝 编辑建议新文件名")
            act_edit.triggered.connect(lambda: self.table.editItem(self.table.item(row, 2)))

            menu.addSeparator()

            # 复制
            act_copy_old = menu.addAction("📋 复制原文件名")
            act_copy_old.triggered.connect(lambda: self._copy_text(old_name))
            if new_name:
                act_copy_new = menu.addAction("📋 复制建议新文件名")
                act_copy_new.triggered.connect(lambda: self._copy_text(new_name))

            menu.addSeparator()

            # 打开文件夹
            act_open = menu.addAction("📂 打开文件所在文件夹")
            act_open.triggered.connect(lambda: self._open_folder(old_name))

            menu.addSeparator()

            # 更改学生
            act_assign = menu.addAction("🔧 选择/更改匹配学生")
            act_assign.triggered.connect(lambda: self._assign_student(row))

            menu.addSeparator()

            # 标记状态
            if status == "✅ 已匹配":
                act_unmatch = menu.addAction("❌ 标记为未匹配")
                act_unmatch.triggered.connect(lambda: self._mark_as_unmatched([row]))
            else:
                act_match = menu.addAction("✅ 标记为已匹配")
                act_match.triggered.connect(lambda: self._mark_as_matched([row]))

        # 批量操作
        if len(rows) > 1:
            act_batch_match = menu.addAction(f"✅ 批量标记为已匹配 ({len(rows)}行)")
            act_batch_match.triggered.connect(lambda: self._mark_as_matched(list(rows)))

            act_batch_unmatch = menu.addAction(f"❌ 批量标记为未匹配 ({len(rows)}行)")
            act_batch_unmatch.triggered.connect(lambda: self._mark_as_unmatched(list(rows)))

        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _copy_text(self, text):
        """复制文本到剪贴板。"""
        QApplication.clipboard().setText(text)
        self.statusBar().showMessage(f"已复制：{text}")

    def _open_folder(self, file_name):
        """在资源管理器中打开并选中文件。"""
        folder = self.le_folder.text().strip()
        full_path = os.path.join(folder, file_name)
        if os.path.exists(full_path):
            subprocess.Popen(["explorer", "/select,", full_path])
        elif os.path.isdir(folder):
            subprocess.Popen(["explorer", folder])

    def _assign_student(self, row):
        """弹出学生列表对话框，选一个学生重新匹配该行。"""
        if not self._name_id_map:
            QMessageBox.warning(self, "提示", "没有可用的学生名单。")
            return

        names = list(self._name_id_map.keys())
        name, ok = QInputDialog.getItem(
            self, "选择学生", "请选择该文件所属的学生：",
            names, editable=False,
        )
        if not ok or not name:
            return

        r = self._results[row]
        student_id = self._name_id_map[name]
        base, ext = os.path.splitext(r["原文件名"])
        template = self.le_template.text().strip() or "{学号}{姓名}{匹配项}"
        extract_keyword = self.le_keyword.text().strip()

        from core.matcher import generate_new_name
        new_name = generate_new_name(template, student_id, name, base, ext, extract_keyword)

        # 更新 _results
        self._results[row]["状态"] = "✅ 已匹配"
        self._results[row]["建议新文件名"] = new_name
        self._results[row]["匹配分数"] = 100
        self._results[row]["匹配方式"] = "手动"

        # 刷新表格
        self._populate_table(self._results)
        self._populate_check_table(self._results)
        self.statusBar().showMessage(f"已指定：{r['原文件名']} → {name}")

    def _mark_as_matched(self, rows):
        """将选中行标记为已匹配（使用原文件名作为建议名）。"""
        for row in rows:
            r = self._results[row]
            if r["状态"] != "✅ 已匹配":
                r["状态"] = "✅ 已匹配"
                r["建议新文件名"] = r["原文件名"]
                r["匹配分数"] = 0
                r["匹配方式"] = "手动"
        self._populate_table(self._results)
        self._populate_check_table(self._results)
        self.statusBar().showMessage(f"已标记 {len(rows)} 行为已匹配")

    def _mark_as_unmatched(self, rows):
        """将选中行标记为未匹配。"""
        for row in rows:
            r = self._results[row]
            r["状态"] = "❌ 未匹配"
            r["建议新文件名"] = ""
            r["匹配分数"] = 0
            r["匹配方式"] = ""
        self._populate_table(self._results)
        self._populate_check_table(self._results)
        self.statusBar().showMessage(f"已标记 {len(rows)} 行为未匹配")

    # ── 异常情况右键菜单 ──────────────────────────────────────────────

    def _on_check_context_menu(self, pos):
        """异常情况表右键菜单：复制名单、导出 CSV。"""
        menu = QMenu(self)

        # 收集缺交和多交名单
        missing_items = []   # "学号姓名"
        missing_names = []
        extra_items   = []   # "学号姓名"
        extra_names   = []
        for row in range(self.check_table.rowCount()):
            status = self.check_table.item(row, 3).text()
            sid  = self.check_table.item(row, 1).text()
            name = self.check_table.item(row, 0).text()
            if status == "❌ 缺交":
                missing_items.append(f"{sid}{name}")
                missing_names.append(name)
            elif status == "⚠️ 多交":
                extra_items.append(f"{sid}{name}")
                extra_names.append(name)

        if missing_items:
            missing_text = "、".join(missing_items)
            act = menu.addAction(f"📋 复制缺交名单（{len(missing_items)}人）")
            act.triggered.connect(lambda _checked=False, t=missing_text: self._copy_text(t))
            names_text = "、".join(missing_names)
            act_name = menu.addAction(f"📋 复制缺交名单-姓名（{len(missing_names)}人）")
            act_name.triggered.connect(lambda _checked=False, t=names_text: self._copy_text(t))

        if extra_items:
            extra_text = "、".join(extra_items)
            menu.addSeparator()
            act = menu.addAction(f"📋 复制多交名单（{len(extra_items)}人）")
            act.triggered.connect(lambda _checked=False, t=extra_text: self._copy_text(t))
            enames_text = "、".join(extra_names)
            act_name2 = menu.addAction(f"📋 复制多交名单-姓名（{len(extra_names)}人）")
            act_name2.triggered.connect(lambda _checked=False, t=enames_text: self._copy_text(t))

        if missing_items or extra_items:
            menu.addSeparator()
            act_export = menu.addAction("📄 导出异常名单 CSV")
            act_export.triggered.connect(self._export_abnormal_csv)

        menu.exec(self.check_table.viewport().mapToGlobal(pos))

    def _export_abnormal_csv(self):
        """导出异常名单（缺交+多交）到 CSV。"""
        import csv
        from datetime import datetime

        rows = []
        for row in range(self.check_table.rowCount()):
            rows.append({
                "姓名": self.check_table.item(row, 0).text(),
                "学号": self.check_table.item(row, 1).text(),
                "提交数": self.check_table.item(row, 2).text(),
                "状态": self.check_table.item(row, 3).text(),
                "匹配文件": self.check_table.item(row, 4).text(),
            })

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(SCRIPT_DIR, f"abnormal_list_{timestamp}.csv")
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=["姓名", "学号", "提交数", "状态", "匹配文件"])
            w.writeheader()
            w.writerows(rows)

        self.statusBar().showMessage(f"异常名单已导出：{path}")

    def _populate_check_table(self, results):
        """将异常情况的异常项（缺交/多交）填入 check_table。"""
        if not self._name_id_map:
            self.check_table.setRowCount(0)
            return

        report = check_submissions(results, self._name_id_map)
        # 过滤掉正常提交，只显示异常；排序：缺交优先 → 多交
        abnormal = [r for r in report if r["状态"] != "✅ 正常"]
        abnormal.sort(key=lambda r: 0 if r["状态"] == "❌ 缺交" else 1)
        self.check_table.setRowCount(len(abnormal))

        for row, r in enumerate(abnormal):
            status = r["状态"]

            item_name    = QTableWidgetItem(r["姓名"])
            item_id      = QTableWidgetItem(r["学号"])
            item_count   = QTableWidgetItem(str(r["提交数"]))
            item_status  = QTableWidgetItem(status)
            item_files   = QTableWidgetItem(r["匹配文件"])

            # 状态颜色
            if status == "❌ 缺交":
                bg = QColor("#ffebee"); fg = QColor("#c62828")
            else:  # 多交
                bg = QColor("#fff8e1"); fg = QColor("#e65100")

            for item in (item_name, item_id, item_count, item_status, item_files):
                item.setBackground(bg)
                item.setForeground(fg)

            self.check_table.setItem(row, 0, item_name)
            self.check_table.setItem(row, 1, item_id)
            self.check_table.setItem(row, 2, item_count)
            self.check_table.setItem(row, 3, item_status)
            self.check_table.setItem(row, 4, item_files)

    # ── 执行重命名 ────────────────────────────────────────────────────

    def _on_apply(self):
        """点击「执行重命名」按钮。"""
        if not self._results:
            QMessageBox.information(self, "提示", "请先点击「预览匹配结果」。")
            return

        matched = [r for r in self._results if r["状态"] == "✅ 已匹配"]
        if not matched:
            QMessageBox.information(self, "提示", "没有可重命名的文件（无 ✅ 已匹配 项）。")
            return

        reply = QMessageBox.question(
            self, "确认重命名",
            f"将对 {len(matched)} 个文件执行重命名，是否继续？",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        self.btn_preview.setEnabled(False)
        self.btn_apply.setEnabled(False)

        # 显示进度条
        self.progress.setVisible(True)
        self.progress.setMaximum(len(matched))
        self.progress.setValue(0)

        # 生成日志路径
        from datetime import datetime
        log_dir = os.path.join(SCRIPT_DIR, "logs")
        os.makedirs(log_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = os.path.join(log_dir, f"rename_log_{timestamp}.csv")

        folder_path = self.le_folder.text().strip()
        self._last_log_path = log_path
        self._rename_worker = RenameWorker(self._results, folder_path, log_path)
        self._rename_worker.progress.connect(self._on_rename_progress)
        self._rename_worker.file_done.connect(self._on_rename_file_done)
        self._rename_worker.all_done.connect(self._on_rename_all_done)
        self._rename_worker.start()

    def _on_rename_progress(self, current, msg):
        self.progress.setValue(current)
        self.statusBar().showMessage(msg)

    def _on_rename_file_done(self, old_name, new_name, success, note):
        """单个文件处理完成时，在表格中标注结果。"""
        for row in range(self.table.rowCount()):
            if self.table.item(row, 1).text() == old_name:
                status_item = self.table.item(row, 0)
                if success:
                    status_item.setText("✔ 已重命名")
                else:
                    status_item.setText(f"⚠️ 跳过 ({note})")
                break

    def _on_rename_all_done(self, renamed, skipped):
        self.progress.setVisible(False)
        self.btn_preview.setEnabled(True)
        self.btn_apply.setEnabled(True)
        self.statusBar().showMessage(
            f"重命名完成 —— ✔ {renamed} 个成功 | ⚠️ {skipped} 个跳过"
        )

    # ── 撤销重命名 ────────────────────────────────────────────────────

    def _on_undo(self):
        """点击「撤销上次重命名」按钮。"""
        if not self._last_log_path:
            QMessageBox.information(self, "提示", "没有可撤销的操作。请先执行一次重命名。")
            return

        if not os.path.exists(self._last_log_path):
            QMessageBox.warning(self, "日志缺失",
                f"日志文件不存在：{self._last_log_path}")
            return

        reply = QMessageBox.question(
            self, "确认撤销",
            f"将根据日志撤销上次重命名操作，是否继续？\n\n日志文件：{self._last_log_path}",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        folder_path = self.le_folder.text().strip()
        undo_renaming(self._last_log_path, folder_path)

        # 撤销后刷新预览
        self.statusBar().showMessage("撤销完成，请重新点击「预览匹配结果」查看当前状态。")
        self._last_log_path = ""


# ═══════════════════════════════════════════════════════════════════════
# 入口
# ═══════════════════════════════════════════════════════════════════════

def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
