"""
学生作业文件批量命名规范化工具 —— 主入口

功能：加载配置 → 校验数据 → 匹配 → 导出 → (可选)执行重命名。
"""

import json
import os
import sys
import argparse
import openpyxl

from core.matcher import build_pinyin_map, match_files
from core.renamer import apply_renaming, undo_renaming
from core.reporter import export_results
from core.checker import check_submissions, print_check_report


def load_config(config_path="config.json"):
    """加载 JSON 配置文件。

    Args:
        config_path: 配置文件路径

    Returns:
        dict: 配置字典
    """
    if not os.path.exists(config_path):
        sys.exit(f"错误：配置文件不存在 —— {config_path}")

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
    except Exception as e:
        sys.exit(f"错误：无法解析配置文件 —— {e}")

    # 必填字段校验
    required = ["folder_path", "excel_path", "template"]
    for key in required:
        if key not in cfg:
            sys.exit(f"错误：config.json 缺少必填字段「{key}」")

    # 设置默认值
    cfg.setdefault("match_threshold", 80)
    cfg.setdefault("enable_pinyin", True)
    cfg.setdefault("output_path", "match_result.csv")

    return cfg


def load_data(folder_path, excel_path):
    """加载并校验输入数据。

    Args:
        folder_path: 待处理文件夹路径
        excel_path: 包含姓名和学号两列的 Excel 文件路径

    Returns:
        (filenames, name_id_map): 文件名列表（含扩展名）和 {姓名: 学号} 映射字典
    """
    # 校验文件夹路径
    if not os.path.exists(folder_path):
        sys.exit(f"错误：文件夹路径不存在 —— {folder_path}")
    if not os.path.isdir(folder_path):
        sys.exit(f"错误：路径不是文件夹 —— {folder_path}")

    # 校验 Excel 文件
    if not os.path.exists(excel_path):
        sys.exit(f"错误：Excel 文件不存在 —— {excel_path}")

    # 根据扩展名选择读取方式（兼容 .xls 和 .xlsx）
    if excel_path.lower().endswith(".xls") and not excel_path.lower().endswith(".xlsx"):
        try:
            import xlrd
            wb = xlrd.open_workbook(excel_path)
            ws = wb.sheet_by_index(0)
            headers = [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
            rows = [
                [ws.cell_value(r, c) for c in range(ws.ncols)]
                for r in range(1, ws.nrows)
            ]
        except Exception as e:
            sys.exit(f"错误：无法读取 .xls 文件 —— {e}")
    else:
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active
            headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = [[c for c in row] for row in ws.iter_rows(min_row=2, values_only=True)]
            wb.close()
        except Exception as e:
            sys.exit(f"错误：无法读取 Excel 文件 —— {e}")

    if "姓名" not in headers or "学号" not in headers:
        sys.exit(
            f"错误：Excel 缺少「姓名」或「学号」列。\n"
            f"当前列名为：{headers}\n"
            f"提示：请使用 GUI 模式（python gui.py）或修改 Excel 列名后重试。"
        )

    col_name = headers.index("姓名")
    col_id   = headers.index("学号")

    # 逐行读取，跳过空值
    name_id_map = {}
    for row in rows:
        name = row[col_name]
        sid  = row[col_id]
        if name and sid:
            name = str(name).strip()
            sid  = str(sid).strip()
            if name and sid:
                name_id_map[name] = sid

    if not name_id_map:
        sys.exit("错误：Excel 中没有有效的姓名-学号数据。")

    # 提取直接子文件（不递归子文件夹）
    filenames = [
        f for f in os.listdir(folder_path)
        if os.path.isfile(os.path.join(folder_path, f))
    ]
    if not filenames:
        sys.exit(f"错误：文件夹为空，未找到任何文件 —— {folder_path}")

    return filenames, name_id_map


def validate_template(template):
    """校验命名模板是否包含非法文件名字符。

    Windows 文件名不允许的字符：\\ / : * ? \" < > |
    """
    illegal_chars = set(r'\/:*?"<>|')
    found = illegal_chars & set(template)
    if found:
        sys.exit(f"错误：模板包含非法字符 —— {''.join(found)}")


def main():
    """程序入口：加载配置 → 校验 → 匹配 → 导出 → (可选)重命名。"""
    # 解析命令行参数
    parser = argparse.ArgumentParser(description="学生作业文件批量命名规范化工具")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="执行实际重命名操作（不加此参数则仅预览）",
    )
    parser.add_argument(
        "--undo",
        default="",
        help="撤销重命名操作，指定日志文件路径（如 logs/rename_log_20260506_193001.csv）",
    )
    parser.add_argument(
        "--config",
        default="config.json",
        help="配置文件路径（默认 config.json）",
    )
    args = parser.parse_args()

    # 加载配置
    cfg = load_config(args.config)

    folder_path = cfg["folder_path"]
    excel_path = cfg["excel_path"]
    output_path = cfg.get("output_path", "match_result.csv")
    template = cfg["template"]
    threshold = cfg["match_threshold"]
    enable_pinyin = cfg["enable_pinyin"]
    extract_keyword = cfg.get("extract_keyword", "")
    log_dir = cfg.get("log_dir", "logs")

    # 校验模板合法性
    validate_template(template)

    print("=" * 50)
    print("学生作业文件批量命名规范化工具")
    print("=" * 50)

    print(f"\n文件夹：{folder_path}")
    print(f"映射表：{excel_path}")
    print(f"命名模板：{template}")
    print(f"匹配阈值：{threshold}")
    print(f"拼音匹配：{'开启' if enable_pinyin else '关闭'}")
    if extract_keyword:
        print(f"提取关键词：{extract_keyword}")

    # 加载数据
    filenames, name_id_map = load_data(folder_path, excel_path)
    print(f"\n已加载 {len(filenames)} 个文件，{len(name_id_map)} 条学生记录")

    # 构建拼音映射表（仅在启用时）
    if enable_pinyin:
        pinyin_map = build_pinyin_map(name_id_map)
    else:
        pinyin_map = {}

    # 执行匹配
    results = match_files(filenames, name_id_map, pinyin_map, template, threshold, enable_pinyin, extract_keyword)

    # 导出结果
    export_results(results, output_path)

    # 提交情况检查
    report = check_submissions(results, name_id_map)
    print_check_report(report)

    # 根据参数决定操作
    if args.undo:
        undo_renaming(args.undo, folder_path)
    elif args.apply:
        timestamp = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = os.path.join(log_dir, f"rename_log_{timestamp}.csv")
        apply_renaming(results, folder_path, log_path)
    else:
        print('\n💡 提示：当前为预览模式。如需执行重命名，请使用 --apply 参数。')
        print('💡 撤销重命名：python main.py --undo logs/rename_log_xxx.csv')


if __name__ == "__main__":
    main()
